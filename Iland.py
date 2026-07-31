import atexit
import csv
import json
import os
import random
import re
import shutil
import tempfile
import time
import zipfile
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import parse_qs, urlparse
from xml.etree import ElementTree as ET

import requests
import undetected_chromedriver as uc
from bs4 import BeautifulSoup
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait

import nopecha_1


TARGET_URL = "https://health.ri.gov/licensing/licensee-lists"
OUTPUT_FILE = "master.csv"
DOWNLOAD_DIR = Path(__file__).resolve().with_name("downloads") / "ri_licensee_lists"
PROXY_FILE = Path(__file__).resolve().with_name("Webshare 10 proxies.txt")

WAIT_SECONDS = 45
PAGE_SETTLE_SECONDS = 1.5
HEADLESS = False
USE_PROXY = False
PROFESSION_KEYWORDS = [
    "Applied Behavior Analyst",
    "Aquatic Venue",
    "Asbestos Program",
    "Dental",
    "Hearing Aid Dealers",
    "Home Nursing Care Provider",
    "Kidney Disease Treat. Center",
    
    "Marriage & Fam. Ther./Mental",
    "Massage",
    "Music Therapy",
    "Naturopathic Medicine",
    "Nursing",
    "Nursing Assistant",
    "Nursing Facility",
    "Nursing Home Administrator",
    "Nursing Service Agency",
    "Occupational Therapy",
    "Pharmacy",
    "Physical Therapy",
    "Physician",
    "Physician Assistant",
    "Podiatric Medicine",
    "Portable x-ray",
    "Psychology",
    "Radiologic Technology",
    "Respiratory Care",
    "Speech Language Path./Audio",
    "X-ray Facility Registrations",
]
 #PROFESSION_KEYWORDS = ["Nursing", "Aquatic venue"]
MATCH_ALIASES = {
    "accupuncture": "acupuncture",
    "acupuncture": "acupuncture",
    "aquaticvenue": "aquaticvenue",
}

STATIC_METADATA_FIELDS = [
    "source_url",
    "profession_keyword",
    "profession_option",
    "license_option",
    "downloaded_file",
    "downloaded_at",
]

_PROXY_EXTENSION_DIRS: List[Path] = []


def normalize_text(text: str) -> str:
    return " ".join((text or "").split()).strip()


def normalize_match_key(text: str) -> str:
    text = normalize_text(text).lower()
    text = re.sub(r"[^a-z0-9]+", "", text)
    return MATCH_ALIASES.get(text, text)


def now_string() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")


def cleanup_proxy_extensions() -> None:
    for path in _PROXY_EXTENSION_DIRS:
        shutil.rmtree(path, ignore_errors=True)


atexit.register(cleanup_proxy_extensions)


def load_proxy_pool() -> List[Dict[str, str]]:
    if not PROXY_FILE.exists():
        return []

    proxies: List[Dict[str, str]] = []
    with PROXY_FILE.open("r", encoding="utf-8-sig") as handle:
        for line_number, raw_line in enumerate(handle, start=1):
            line = normalize_text(raw_line)
            if not line or line.startswith("#"):
                continue
            parts = [part.strip() for part in line.split(":")]
            if len(parts) != 4:
                raise ValueError(
                    f"Invalid proxy entry on line {line_number} in {PROXY_FILE.name}: {raw_line.strip()!r}"
                )
            host, port, username, password = parts
            proxies.append(
                {
                    "host": host,
                    "port": port,
                    "username": username,
                    "password": password,
                }
            )
    return proxies


def build_proxy_extension(proxy: Dict[str, str]) -> str:
    extension_dir = Path(tempfile.mkdtemp(prefix="webshare_proxy_"))
    manifest = {
        "manifest_version": 3,
        "name": "Webshare Auth Proxy",
        "version": "1.0.0",
        "permissions": ["webRequest", "webRequestAuthProvider"],
        "host_permissions": ["<all_urls>"],
        "background": {"service_worker": "background.js"},
    }
    background_js = f"""
const proxyCredentials = {{
  username: {json.dumps(proxy["username"])},
  password: {json.dumps(proxy["password"])},
}};

chrome.webRequest.onAuthRequired.addListener(
  (details, callback) => {{
    callback({{authCredentials: proxyCredentials}});
  }},
  {{urls: ["<all_urls>"]}},
  ["asyncBlocking"]
);
""".strip()

    (extension_dir / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    (extension_dir / "background.js").write_text(background_js, encoding="utf-8")
    _PROXY_EXTENSION_DIRS.append(extension_dir)
    return str(extension_dir)


def build_driver() -> Tuple[uc.Chrome, Optional[Dict[str, str]]]:
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    options = uc.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-background-networking")
    options.add_experimental_option(
        "prefs",
        {
            "download.default_directory": str(DOWNLOAD_DIR),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
            "profile.default_content_setting_values.automatic_downloads": 1,
        },
    )

    proxy = None
    if USE_PROXY:
        proxies = load_proxy_pool()
        if not proxies:
            raise RuntimeError(f"No valid proxies found in {PROXY_FILE}")
        proxy = random.choice(proxies)
        proxy_url = f"http://{proxy['host']}:{proxy['port']}"
        extension_dir = build_proxy_extension(proxy)
        options.add_argument(f"--proxy-server={proxy_url}")
        options.add_argument(f"--disable-extensions-except={extension_dir}")
        options.add_argument(f"--load-extension={extension_dir}")
        print(f"[proxy] Using {proxy['host']}:{proxy['port']}")

    if HEADLESS:
        options.add_argument("--headless=new")

    driver = uc.Chrome(options=options, version_main=150, use_subprocess=True)
    driver.set_page_load_timeout(120)
    driver.set_script_timeout(60)
    driver.execute_cdp_cmd(
        "Page.setDownloadBehavior",
        {"behavior": "allow", "downloadPath": str(DOWNLOAD_DIR)},
    )
    return driver, proxy


def wait_for_page_ready(driver, wait: WebDriverWait) -> None:
    wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
    time.sleep(PAGE_SETTLE_SECONDS)


def wait_for_element_present(
    driver,
    by: By,
    value: str,
    timeout: Optional[int] = None,
    poll_interval: float = 1.0,
) -> bool:
    deadline = None if timeout is None else time.time() + timeout
    while deadline is None or time.time() < deadline:
        try:
            if driver.find_elements(by, value):
                return True
        except WebDriverException:
            pass
        time.sleep(poll_interval)
    return False


def find_frame_with_element(driver, by, value) -> bool:
    """Search the top document and every iframe for an element.

    Leaves the driver switched into whichever frame contains the element
    (or in the default content if it lives in the top document). Returns
    True when the element is found so subsequent Select/click calls run in
    the correct frame context.
    """
    driver.switch_to.default_content()
    if driver.find_elements(by, value):
        return True

    frame_count = len(driver.find_elements(By.TAG_NAME, "iframe"))
    for index in range(frame_count):
        try:
            driver.switch_to.default_content()
            frames = driver.find_elements(By.TAG_NAME, "iframe")
            if index >= len(frames):
                break
            driver.switch_to.frame(frames[index])
            if driver.find_elements(by, value):
                return True
        except WebDriverException:
            continue

    driver.switch_to.default_content()
    return False


def wait_for_element_in_any_frame(
    driver,
    by,
    value: str,
    timeout: Optional[int] = None,
    poll_interval: float = 1.0,
) -> bool:
    deadline = None if timeout is None else time.time() + timeout
    while deadline is None or time.time() < deadline:
        try:
            if find_frame_with_element(driver, by, value):
                return True
        except WebDriverException:
            pass
        time.sleep(poll_interval)
    return False


def describe_iframes(driver) -> None:
    try:
        driver.switch_to.default_content()
        frames = driver.find_elements(By.TAG_NAME, "iframe")
    except WebDriverException:
        return
    if not frames:
        print("[page] No iframes found in the top document.")
        return
    print(f"[page] Found {len(frames)} iframe(s):")
    for index, frame in enumerate(frames):
        try:
            src = frame.get_attribute("src") or "(no src)"
        except WebDriverException:
            src = "(unavailable)"
        print(f"        iframe[{index}] src={src}")


def is_cloudflare_challenge_page(driver) -> bool:
    title = normalize_text(driver.title).lower()
    if "just a moment" in title:
        return True
    html = driver.page_source.lower()
    markers = [
        "cf-challenge",
        "cf-turnstile",
        "challenge-platform",
        "/cdn-cgi/challenge-platform/",
    ]
    return any(marker in html for marker in markers)


def get_browser_cookies_for_nopecha(driver) -> List[Dict[str, object]]:
    cookies = []
    for item in driver.get_cookies():
        cookie = {
            "name": item.get("name", ""),
            "value": item.get("value", ""),
            "domain": item.get("domain", ""),
            "path": item.get("path", "/"),
            "hostOnly": not str(item.get("domain", "")).startswith("."),
            "httpOnly": bool(item.get("httpOnly")),
            "secure": bool(item.get("secure")),
            "session": item.get("expiry") is None,
        }
        if item.get("expiry") is not None:
            cookie["expirationDate"] = int(item["expiry"])
        cookies.append(cookie)
    return cookies


def parse_turnstile_from_iframe_src(src: str) -> Dict[str, str]:
    if not src:
        return {}
    parsed = urlparse(src)
    query = parse_qs(parsed.query)
    result = {}
    if "k" in query and query["k"]:
        result["sitekey"] = query["k"][0]
    if "action" in query and query["action"]:
        result["action"] = query["action"][0]
    if "cData" in query and query["cData"]:
        result["cdata"] = query["cData"][0]
    return result


def get_turnstile_context(driver) -> Optional[Dict[str, str]]:
    script = """
    const widget =
      document.querySelector('[data-sitekey][class*="turnstile"]') ||
      document.querySelector('[data-sitekey][id*="turnstile"]') ||
      document.querySelector('[data-sitekey]');
    if (widget) {
      return {
        sitekey: widget.getAttribute('data-sitekey') || '',
        action: widget.getAttribute('data-action') || '',
        cdata: widget.getAttribute('data-cdata') || widget.getAttribute('data-cData') || ''
      };
    }
    const iframe = document.querySelector('iframe[src*="turnstile"], iframe[src*="challenge-platform"]');
    if (iframe) {
      return {
        sitekey: '',
        action: iframe.getAttribute('data-action') || '',
        cdata: iframe.getAttribute('data-cdata') || ''
      };
    }
    return null;
    """
    try:
        data = driver.execute_script(script)
    except WebDriverException:
        data = None

    data = data or {}
    iframe_data = {}
    try:
        iframe = driver.find_element(By.CSS_SELECTOR, 'iframe[src*="turnstile"], iframe[src*="challenge-platform"]')
        iframe_data = parse_turnstile_from_iframe_src(iframe.get_attribute("src") or "")
    except NoSuchElementException:
        pass

    merged = {
        "sitekey": data.get("sitekey") or iframe_data.get("sitekey") or "",
        "action": data.get("action") or iframe_data.get("action") or "",
        "cdata": data.get("cdata") or iframe_data.get("cdata") or "",
    }
    return merged if merged.get("sitekey") else None


def solve_turnstile_via_nopecha(driver, proxy: Optional[Dict[str, str]]) -> bool:
    context = get_turnstile_context(driver)
    if not context:
        return False
    if proxy is None:
        print("[captcha] Turnstile detected but no proxy is configured for NopeCHA.")
        return False

    payload = {
        "sitekey": context["sitekey"],
        "url": driver.current_url,
        "proxy": {
            "scheme": "http",
            "host": proxy["host"],
            "port": proxy["port"],
            "username": proxy["username"],
            "password": proxy["password"],
        },
        "cookie": get_browser_cookies_for_nopecha(driver),
        "useragent": driver.execute_script("return navigator.userAgent"),
    }
    data = {key: context[key] for key in ("action", "cdata") if context.get(key)}
    if data:
        payload["data"] = data

    print("[captcha] Solving Cloudflare Turnstile via NopeCHA...")
    response = nopecha_1.SESSION.post(
        f"{nopecha_1.BASE_URL}/token/turnstile",
        headers=nopecha_1.HEADERS,
        json=payload,
        timeout=30,
    )
    if response.status_code >= 400:
        raise nopecha_1.NopechaError(
            f"submit token/turnstile -> HTTP {response.status_code}: {response.text[:200]}"
        )
    job_id = (response.json() or {}).get("data")
    if not job_id:
        raise nopecha_1.NopechaError(
            f"submit token/turnstile returned no job id: {response.text[:200]}"
        )

    token = nopecha_1._poll("token/turnstile", job_id, nopecha_1.DEFAULT_TIMEOUT, nopecha_1.POLL_INTERVAL)
    if isinstance(token, list):
        token = token[0] if token else ""
    token = str(token).strip()
    if not token:
        return False

    inject_script = """
    const token = arguments[0];
    let updated = 0;
    for (const el of document.querySelectorAll(
      'input[name="cf-turnstile-response"], textarea[name="cf-turnstile-response"], ' +
      'input[name="g-recaptcha-response"], textarea[name="g-recaptcha-response"]'
    )) {
      el.value = token;
      el.innerHTML = token;
      el.dispatchEvent(new Event('input', {bubbles: true}));
      el.dispatchEvent(new Event('change', {bubbles: true}));
      updated += 1;
    }

    const seen = new Set();
    let callbacksCalled = 0;
    const walk = (obj, depth) => {
      if (!obj || depth > 5 || typeof obj !== 'object') return;
      if (seen.has(obj)) return;
      seen.add(obj);
      for (const key of Object.keys(obj)) {
        let value;
        try { value = obj[key]; } catch (e) { continue; }
        if (typeof value === 'function' && /callback/i.test(key)) {
          try {
            value(token);
            callbacksCalled += 1;
          } catch (e) {}
        } else if (value && typeof value === 'object') {
          walk(value, depth + 1);
        }
      }
    };
    walk(window, 0);

    const form =
      document.querySelector('form#challenge-form') ||
      document.querySelector('form[action*="challenge"]');
    if (form) {
      form.dispatchEvent(new Event('submit', {bubbles: true, cancelable: true}));
    }
    return {updated, callbacksCalled};
    """
    driver.execute_script(inject_script, token)
    time.sleep(5)
    return True


def open_target_page(driver, wait: WebDriverWait, proxy: Optional[Dict[str, str]]) -> None:
    driver.get(TARGET_URL)
    wait_for_page_ready(driver, wait)

    if wait_for_element_in_any_frame(driver, By.ID, "jx-profession", timeout=5, poll_interval=0.5):
        return

    if is_cloudflare_challenge_page(driver):
        print("[captcha] Cloudflare challenge detected.")
    else:
        print("[page] Waiting for the profession dropdown to appear.")
        describe_iframes(driver)

    print(
        "[captcha] Solve Cloudflare manually in the browser first. "
        "The script will keep waiting and continue automatically once the profession dropdown appears."
    )
    if wait_for_element_in_any_frame(driver, By.ID, "jx-profession", timeout=None, poll_interval=2.0):
        print("[page] Profession dropdown detected. Starting automation.")
        return


def find_matching_professions(driver, wait: WebDriverWait) -> List[Tuple[str, str, str]]:
    select = Select(wait.until(EC.presence_of_element_located((By.ID, "jx-profession"))))
    wanted = {normalize_match_key(keyword): keyword for keyword in PROFESSION_KEYWORDS}
    matches: List[Tuple[str, str, str]] = []
    seen = set()

    for option in select.options:
        value = option.get_attribute("value") or ""
        text = normalize_text(option.text)
        if not value or not text:
            continue
        option_key = normalize_match_key(text)
        if option_key in wanted and option_key not in seen:
            matches.append((value, text, wanted[option_key]))
            seen.add(option_key)
    return matches


def select_profession(driver, wait: WebDriverWait, profession_value: str) -> None:
    profession_select = Select(wait.until(EC.presence_of_element_located((By.ID, "jx-profession"))))
    old_license_options = [normalize_text(opt.text) for opt in Select(driver.find_element(By.ID, "jx-license")).options]
    profession_select.select_by_value(profession_value)
    time.sleep(1)

    def license_options_changed(d) -> bool:
        try:
            options = [normalize_text(opt.text) for opt in Select(d.find_element(By.ID, "jx-license")).options]
            return options != old_license_options or len(options) > 1
        except Exception:
            return False

    wait.until(license_options_changed)
    time.sleep(PAGE_SETTLE_SECONDS)


def get_license_options(driver, wait: WebDriverWait) -> List[Tuple[str, str]]:
    license_select = Select(wait.until(EC.presence_of_element_located((By.ID, "jx-license"))))
    options = []
    for option in license_select.options:
        value = option.get_attribute("value") or ""
        text = normalize_text(option.text)
        if not value or not text:
            continue
        options.append((value, text))
    return options


def click_element(driver, element) -> None:
    try:
        element.click()
    except WebDriverException:
        driver.execute_script("arguments[0].click();", element)


def wait_for_download_button(driver, wait: WebDriverWait):
    return wait.until(EC.element_to_be_clickable((By.ID, "downloadbutton")))


def trigger_search(driver, wait: WebDriverWait) -> None:
    button = wait.until(EC.element_to_be_clickable((By.ID, "countbutton")))
    click_element(driver, button)
    time.sleep(2)
    wait_for_download_button(driver, wait)
    time.sleep(PAGE_SETTLE_SECONDS)


def list_download_files() -> Dict[str, float]:
    files = {}
    for path in DOWNLOAD_DIR.glob("*"):
        if path.is_file():
            files[path.name] = path.stat().st_mtime
    return files


def wait_for_new_download(before: Dict[str, float], timeout: int = 180) -> Path:
    deadline = time.time() + timeout
    newest_candidate = None
    while time.time() < deadline:
        active_partial = False
        for path in DOWNLOAD_DIR.glob("*"):
            if not path.is_file():
                continue
            if path.suffix.lower() in {".crdownload", ".tmp", ".part"}:
                active_partial = True
                continue
            mtime = path.stat().st_mtime
            if path.name not in before or mtime > before[path.name] + 0.5:
                newest_candidate = path
        if newest_candidate and not active_partial:
            return newest_candidate
        time.sleep(1)
    raise TimeoutException("Timed out waiting for downloaded file")


def download_license_file(driver, wait: WebDriverWait) -> Path:
    before = list_download_files()
    button = wait_for_download_button(driver, wait)
    click_element(driver, button)
    return wait_for_new_download(before)


def clean_header(value) -> str:
    text = normalize_text("" if value is None else str(value))
    return text or "column"


def coerce_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return normalize_text(str(value))


def rows_from_csv(path: Path) -> List[Dict[str, str]]:
    rows = []
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            rows.append({clean_header(key): coerce_cell(value) for key, value in row.items()})
    return rows


def rows_from_xlsx_openpyxl(path: Path) -> List[Dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [clean_header(value) for value in next(iterator)]
    rows = []
    for raw_row in iterator:
        row = {}
        for index, header in enumerate(headers):
            row[header] = coerce_cell(raw_row[index] if index < len(raw_row) else "")
        if any(row.values()):
            rows.append(row)
    workbook.close()
    return rows


def rows_from_html_table(path: Path) -> List[Dict[str, str]]:
    soup = BeautifulSoup(path.read_text(encoding="utf-8", errors="ignore"), "html.parser")
    table = soup.find("table")
    if table is None:
        return []

    rows = []
    header_cells = table.find("tr")
    if header_cells is None:
        return []
    headers = [clean_header(cell.get_text(" ", strip=True)) for cell in header_cells.find_all(["th", "td"])]
    for tr in table.find_all("tr")[1:]:
        cells = tr.find_all(["td", "th"])
        if not cells:
            continue
        row = {}
        for index, header in enumerate(headers):
            text = cells[index].get_text(" ", strip=True) if index < len(cells) else ""
            row[header] = normalize_text(text)
        if any(row.values()):
            rows.append(row)
    return rows


def rows_from_xlsx_xml(path: Path) -> List[Dict[str, str]]:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    shared_strings: List[str] = []
    rows: List[List[str]] = []

    with zipfile.ZipFile(path) as archive:
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for si in root.findall("main:si", ns):
                text_parts = [t.text or "" for t in si.findall(".//main:t", ns)]
                shared_strings.append("".join(text_parts))

        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        first_sheet = workbook.find("main:sheets/main:sheet", ns)
        if first_sheet is None:
            return []
        rel_id = first_sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")

        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        target = None
        for rel in rels.findall("pkg:Relationship", ns):
            if rel.attrib.get("Id") == rel_id:
                target = rel.attrib.get("Target")
                break
        if not target:
            return []
        target_path = "xl/" + target.lstrip("/")
        root = ET.fromstring(archive.read(target_path))
        for row_el in root.findall(".//main:sheetData/main:row", ns):
            values: List[str] = []
            for cell in row_el.findall("main:c", ns):
                cell_type = cell.attrib.get("t")
                value_el = cell.find("main:v", ns)
                value = value_el.text if value_el is not None else ""
                if cell_type == "s" and value.isdigit():
                    index = int(value)
                    value = shared_strings[index] if index < len(shared_strings) else ""
                values.append(normalize_text(value))
            rows.append(values)

    if not rows:
        return []
    headers = [clean_header(value) for value in rows[0]]
    parsed_rows = []
    for raw_row in rows[1:]:
        row = {}
        for index, header in enumerate(headers):
            row[header] = raw_row[index] if index < len(raw_row) else ""
        if any(row.values()):
            parsed_rows.append(row)
    return parsed_rows


def load_rows_from_download(path: Path) -> List[Dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return rows_from_csv(path)

    if suffix == ".xlsx":
        try:
            return rows_from_xlsx_openpyxl(path)
        except Exception:
            return rows_from_xlsx_xml(path)

    if suffix in {".xls", ".html", ".htm"}:
        html_rows = rows_from_html_table(path)
        if html_rows:
            return html_rows
        try:
            import pandas as pd

            frame = pd.read_excel(path, dtype=str)
            return [
                {clean_header(key): coerce_cell(value) for key, value in row.items()}
                for row in frame.fillna("").to_dict(orient="records")
            ]
        except Exception:
            return []

    try:
        return rows_from_csv(path)
    except Exception:
        return []


def read_existing_master(path: Path) -> Tuple[List[Dict[str, str]], List[str]]:
    if not path.exists() or path.stat().st_size == 0:
        return [], []
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        headers = reader.fieldnames or []
        rows = [{clean_header(key): coerce_cell(value) for key, value in row.items()} for row in reader]
    return rows, headers


def dedupe_rows(rows: Iterable[Dict[str, str]], headers: Sequence[str]) -> List[Dict[str, str]]:
    seen = set()
    unique_rows = []
    for row in rows:
        key = tuple(row.get(header, "") for header in headers)
        if key in seen:
            continue
        seen.add(key)
        unique_rows.append(row)
    return unique_rows


def write_master_csv(path: Path, rows: List[Dict[str, str]]) -> None:
    headers = list(STATIC_METADATA_FIELDS)
    dynamic_headers = sorted(
        {
            key
            for row in rows
            for key in row.keys()
            if key not in STATIC_METADATA_FIELDS
        }
    )
    headers.extend(dynamic_headers)
    final_rows = dedupe_rows(rows, headers)

    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with tmp_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in final_rows:
            writer.writerow({header: row.get(header, "") for header in headers})
    os.replace(tmp_path, path)


def merge_download_into_master(
    master_path: Path,
    downloaded_path: Path,
    profession_keyword: str,
    profession_text: str,
    license_text: str,
) -> int:
    source_rows = load_rows_from_download(downloaded_path)
    if not source_rows:
        return 0

    existing_rows, _ = read_existing_master(master_path)
    downloaded_at = now_string()
    for row in source_rows:
        row.update(
            {
                "source_url": TARGET_URL,
                "profession_keyword": profession_keyword,
                "profession_option": profession_text,
                "license_option": license_text,
                "downloaded_file": downloaded_path.name,
                "downloaded_at": downloaded_at,
            }
        )
    all_rows = existing_rows + source_rows
    write_master_csv(master_path, all_rows)
    return len(source_rows)


def main() -> None:
    master_path = Path(__file__).resolve().with_name(OUTPUT_FILE)
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    driver = None
    try:
        driver, proxy = build_driver()
        wait = WebDriverWait(driver, WAIT_SECONDS)
        open_target_page(driver, wait, proxy)

        professions = find_matching_professions(driver, wait)
        if not professions:
            raise RuntimeError(
                f"No matching profession options found for {PROFESSION_KEYWORDS!r} on {TARGET_URL}"
            )

        total_rows = 0
        total_files = 0

        for profession_value, profession_text, profession_keyword in professions:
            print(f"[profession] {profession_text}")
            select_profession(driver, wait, profession_value)
            license_options = get_license_options(driver, wait)
            if not license_options:
                print(f"  [skip] No license options loaded for {profession_text}")
                continue

            for license_value, license_text in license_options:
                print(f"  [license] {license_text}")
                select_profession(driver, wait, profession_value)
                license_select = Select(wait.until(EC.presence_of_element_located((By.ID, "jx-license"))))
                license_select.select_by_value(license_value)
                time.sleep(PAGE_SETTLE_SECONDS)

                trigger_search(driver, wait)
                downloaded_path = download_license_file(driver, wait)
                rows_added = merge_download_into_master(
                    master_path,
                    downloaded_path,
                    profession_keyword,
                    profession_text,
                    license_text,
                )
                total_files += 1
                total_rows += rows_added
                print(f"    [saved] {downloaded_path.name} -> {rows_added} rows into {master_path.name}")

        print(f"[done] {total_files} files merged, {total_rows} rows written to {master_path.name}")
    finally:
        if driver is not None:
            try:
                driver.quit()
            except Exception:
                pass


if __name__ == "__main__":
    main()
