import csv
import os
import time
import zipfile
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple
from xml.etree import ElementTree as ET

import undetected_chromedriver as uc
from bs4 import BeautifulSoup
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait


TARGET_URL = "https://ebizws.mt.gov/PUBLICPORTAL/searchform?mylist=licenses"
TARGET_BOARDS = [
    "Board of Dentistry",
    "Board of Massage Therapy",
    "Board of Nursing Home Administrators",
]
OUTPUT_FILE = "Master_Monota.csv"
DOWNLOAD_DIR = Path(__file__).resolve().with_name("downloads") / "monota"

WAIT_SECONDS = 60
PAGE_SETTLE_SECONDS = 1.5
HEADLESS = False

STATIC_METADATA_FIELDS = [
    "source_url",
    "board_name",
    "downloaded_file",
    "downloaded_at",
]


def normalize_text(text: str) -> str:
    return " ".join((text or "").split()).strip()


def now_string() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")


def build_driver():
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    options = uc.ChromeOptions()
    options.add_argument("--start-maximized")
    # options.add_argument("--disable-blink-features=AutomationControlled")
    # options.add_argument("--disable-notifications")
    # options.add_argument("--disable-popup-blocking")
    # options.add_argument("--no-first-run")
    # options.add_argument("--no-default-browser-check")
    # options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option(
        "prefs",
        {
            "download.default_directory": str(DOWNLOAD_DIR),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
            "profile.default_content_setting_values.automatic_downloads": 1,
        },
    )

    if HEADLESS:
        options.add_argument("--headless=new")

    driver = uc.Chrome(options=options, version_main=150, use_subprocess=True)
    driver.set_page_load_timeout(120)
    driver.set_script_timeout(60)
    driver.execute_cdp_cmd(
        "Page.setDownloadBehavior",
        {"behavior": "allow", "downloadPath": str(DOWNLOAD_DIR)},
    )
    return driver


def wait_for_page_ready(driver, wait: WebDriverWait) -> None:
    wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
    time.sleep(PAGE_SETTLE_SECONDS)


def click_element(driver, element) -> None:
    try:
        element.click()
    except WebDriverException:
        driver.execute_script("arguments[0].click();", element)


def open_search_form(driver, wait: WebDriverWait) -> None:
    driver.get(TARGET_URL)
    wait_for_page_ready(driver, wait)
    wait.until(EC.presence_of_element_located((By.ID, "licboard")))


def select_board(driver, wait: WebDriverWait, board_name: str) -> None:
    select = Select(wait.until(EC.element_to_be_clickable((By.ID, "licboard"))))
    available = [normalize_text(option.text) for option in select.options if normalize_text(option.text)]
    if board_name not in available:
        raise RuntimeError(f"Board {board_name!r} not found. Available options: {available}")
    select.select_by_visible_text(board_name)
    time.sleep(PAGE_SETTLE_SECONDS)


def wait_for_user_verification(board_name: str) -> None:
    input(
        f"Solve any verification for '{board_name}' in the open browser, "
        "then press Enter here to continue with Search..."
    )


def submit_search(driver, wait: WebDriverWait) -> None:
    search_button = wait.until(EC.element_to_be_clickable((By.ID, "submitbtn")))
    click_element(driver, search_button)
    wait.until(EC.presence_of_element_located((By.ID, "csvexport")))
    wait.until(EC.element_to_be_clickable((By.ID, "csvexport")))
    time.sleep(PAGE_SETTLE_SECONDS)


def list_download_files() -> Dict[str, float]:
    files = {}
    for path in DOWNLOAD_DIR.glob("*"):
        if path.is_file():
            files[path.name] = path.stat().st_mtime
    return files


def wait_for_new_download(before: Dict[str, float], timeout: int = 180) -> Path:
    deadline = time.time() + timeout
    newest_candidate = None
    while time.time() < deadline:
        active_partial = False
        for path in DOWNLOAD_DIR.glob("*"):
            if not path.is_file():
                continue
            if path.suffix.lower() in {".crdownload", ".tmp", ".part"}:
                active_partial = True
                continue
            mtime = path.stat().st_mtime
            if path.name not in before or mtime > before[path.name] + 0.5:
                newest_candidate = path
        if newest_candidate and not active_partial:
            return newest_candidate
        time.sleep(1)
    raise TimeoutException("Timed out waiting for downloaded file")


def download_results_file(driver, wait: WebDriverWait) -> Path:
    before = list_download_files()
    download_button = wait.until(EC.element_to_be_clickable((By.ID, "csvexport")))
    click_element(driver, download_button)
    return wait_for_new_download(before)


def clean_header(value) -> str:
    text = normalize_text("" if value is None else str(value))
    return text or "column"


def coerce_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return normalize_text(str(value))


def rows_from_csv(path: Path) -> List[Dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return [{clean_header(key): coerce_cell(value) for key, value in row.items()} for row in reader]


def rows_from_html_table(path: Path) -> List[Dict[str, str]]:
    soup = BeautifulSoup(path.read_text(encoding="utf-8", errors="ignore"), "html.parser")
    table = soup.find("table")
    if table is None:
        return []

    header_row = table.find("tr")
    if header_row is None:
        return []

    headers = [clean_header(cell.get_text(" ", strip=True)) for cell in header_row.find_all(["th", "td"])]
    rows = []
    for tr in table.find_all("tr")[1:]:
        cells = tr.find_all(["td", "th"])
        if not cells:
            continue
        row = {}
        for index, header in enumerate(headers):
            row[header] = normalize_text(cells[index].get_text(" ", strip=True)) if index < len(cells) else ""
        if any(row.values()):
            rows.append(row)
    return rows


def rows_from_xlsx_openpyxl(path: Path) -> List[Dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [clean_header(value) for value in next(iterator)]
    rows = []
    for raw_row in iterator:
        row = {}
        for index, header in enumerate(headers):
            row[header] = coerce_cell(raw_row[index] if index < len(raw_row) else "")
        if any(row.values()):
            rows.append(row)
    workbook.close()
    return rows


def rows_from_xlsx_xml(path: Path) -> List[Dict[str, str]]:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    shared_strings: List[str] = []
    rows: List[List[str]] = []

    with zipfile.ZipFile(path) as archive:
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for si in root.findall("main:si", ns):
                text_parts = [t.text or "" for t in si.findall(".//main:t", ns)]
                shared_strings.append("".join(text_parts))

        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        first_sheet = workbook.find("main:sheets/main:sheet", ns)
        if first_sheet is None:
            return []

        rel_id = first_sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

        target = None
        for rel in rels.findall("pkg:Relationship", ns):
            if rel.attrib.get("Id") == rel_id:
                target = rel.attrib.get("Target")
                break

        if not target:
            return []

        root = ET.fromstring(archive.read("xl/" + target.lstrip("/")))
        for row_el in root.findall(".//main:sheetData/main:row", ns):
            values: List[str] = []
            for cell in row_el.findall("main:c", ns):
                cell_type = cell.attrib.get("t")
                value_el = cell.find("main:v", ns)
                value = value_el.text if value_el is not None else ""
                if cell_type == "s" and value.isdigit():
                    index = int(value)
                    value = shared_strings[index] if index < len(shared_strings) else ""
                values.append(normalize_text(value))
            rows.append(values)

    if not rows:
        return []

    headers = [clean_header(value) for value in rows[0]]
    parsed_rows = []
    for raw_row in rows[1:]:
        row = {}
        for index, header in enumerate(headers):
            row[header] = raw_row[index] if index < len(raw_row) else ""
        if any(row.values()):
            parsed_rows.append(row)
    return parsed_rows


def load_rows_from_download(path: Path) -> List[Dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return rows_from_csv(path)
    if suffix == ".xlsx":
        try:
            return rows_from_xlsx_openpyxl(path)
        except Exception:
            return rows_from_xlsx_xml(path)
    if suffix in {".xls", ".html", ".htm"}:
        return rows_from_html_table(path)
    try:
        return rows_from_csv(path)
    except Exception:
        return []


def read_existing_master(path: Path) -> Tuple[List[Dict[str, str]], List[str]]:
    if not path.exists() or path.stat().st_size == 0:
        return [], []
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        headers = reader.fieldnames or []
        rows = [{clean_header(key): coerce_cell(value) for key, value in row.items()} for row in reader]
    return rows, headers


def dedupe_rows(rows: Iterable[Dict[str, str]], headers: Sequence[str]) -> List[Dict[str, str]]:
    seen = set()
    unique_rows = []
    for row in rows:
        key = tuple(row.get(header, "") for header in headers)
        if key in seen:
            continue
        seen.add(key)
        unique_rows.append(row)
    return unique_rows


def write_master_csv(path: Path, rows: List[Dict[str, str]]) -> None:
    headers = list(STATIC_METADATA_FIELDS)
    dynamic_headers = sorted({key for row in rows for key in row.keys() if key not in STATIC_METADATA_FIELDS})
    headers.extend(dynamic_headers)
    final_rows = dedupe_rows(rows, headers)

    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with tmp_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in final_rows:
            writer.writerow({header: row.get(header, "") for header in headers})
    os.replace(tmp_path, path)


def merge_download_into_master(master_path: Path, downloaded_path: Path, board_name: str) -> int:
    source_rows = load_rows_from_download(downloaded_path)
    if not source_rows:
        return 0

    existing_rows, _ = read_existing_master(master_path)
    downloaded_at = now_string()
    for row in source_rows:
        row.update(
            {
                "source_url": TARGET_URL,
                "board_name": board_name,
                "downloaded_file": downloaded_path.name,
                "downloaded_at": downloaded_at,
            }
        )

    all_rows = existing_rows + source_rows
    write_master_csv(master_path, all_rows)
    return len(source_rows)


def main() -> None:
    master_path = Path(__file__).resolve().with_name(OUTPUT_FILE)
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    driver = None
    try:
        driver = build_driver()
        wait = WebDriverWait(driver, WAIT_SECONDS)

        total_files = 0
        total_rows = 0

        for board_name in TARGET_BOARDS:
            print(f"[board] {board_name}")
            open_search_form(driver, wait)
            select_board(driver, wait, board_name)
            wait_for_user_verification(board_name)
            submit_search(driver, wait)
            downloaded_path = download_results_file(driver, wait)
            rows_added = merge_download_into_master(master_path, downloaded_path, board_name)
            total_files += 1
            total_rows += rows_added
            print(f"  [saved] {downloaded_path.name} -> {rows_added} rows into {master_path.name}")

        print(f"[done] {total_files} files merged, {total_rows} rows written to {master_path.name}")
    finally:
        if driver is not None:
            try:
                driver.quit()
            except Exception:
                pass


if __name__ == "__main__":
    main()
